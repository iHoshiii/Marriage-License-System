import openpyxl
from openpyxl.drawing.image import Image
from datetime import datetime
import sys
import json
import os
import io
import warnings

# Suppress potential library warnings that can corrupt the binary stream
warnings.filterwarnings("ignore", category=UserWarning)

def cm_to_pixels(cm):
    return int((cm / 2.54) * 96)
def process_excel(data):
    try:
        # Load paths relative to script location
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, "data", "APPLICATION-for-MARRIAGE-LICENSE.xlsx")

        # Use provided image path or skip image insertion
        img_path = data.get("coupleImagePath")
        
        if not os.path.exists(template_path):
            sys.stderr.write(f"Error: Template not found at {template_path}\n")
            sys.stderr.flush()
            os._exit(1)

        wb = openpyxl.load_workbook(template_path)
        
        # Date Calculations
        now = datetime.now()
        day_now = str(now.day)
        month_now = now.strftime("%B")
        year_now = str(now.year)

        # Helper to safely get integers
        def get_int(key):
            val = data.get(key, 0)
            try:
                return int(val) if val else 0   
            except:
                return 0

        # DATA EXTRACTION
        g_age = get_int("gAge")
        b_age = get_int("bAge")
        
        g_town_prov = (f"{data.get('gTown', '')}, {data.get('gProv', 'Nueva Vizcaya')}")
        b_town_prov = (f"{data.get('bTown', '')}, {data.get('bProv', 'Nueva Vizcaya')}")
        g_full_addr = (f"Brgy. , {data.get('gBrgy', '')}, {g_town_prov}")
        b_full_addr = (f"Brgy. , {data.get('bBrgy', '')}, {b_town_prov}")

        # Image Logic
        if img_path and "Notice" in wb.sheetnames:
            if os.path.exists(img_path):
                try:
                    couple_img = Image(img_path)
                    couple_img.height = cm_to_pixels(3.75)
                    couple_img.width = cm_to_pixels(5.73)
                    wb["Notice"].add_image(couple_img, "T11")
                    sys.stderr.write(f"Successfully added image from: {img_path}\n")
                    sys.stderr.flush()
                except Exception as e:
                    sys.stderr.write(f"Warning: Image overlay failed: {e}\n")
                    sys.stderr.flush()
            else:
                sys.stderr.write(f"Error: Image path does not exist: {img_path}\n")
                sys.stderr.flush()
        else:
            sys.stderr.write(f"Notice: Image processing skipped (img_path: {img_path}, sheet: {'Notice' in wb.sheetnames})\n")
            sys.stderr.flush()

        # Set Notice as active sheet if it exists
        if "Notice" in wb.sheetnames:
            wb.active = wb.sheetnames.index("Notice")
            sys.stderr.write("Set 'Notice' as active sheet\n")
            sys.stderr.flush()

        # MAIN APPLICATION SHEET MAPPING
        if "APPLICATION" not in wb.sheetnames:
            sys.stderr.write("Error: 'APPLICATION' sheet missing in template\n")
            os._exit(1)
            
        app = wb["APPLICATION"]
        
        # Groom
        app['B8'], app['B9'], app['B10'] =(data.get("gFirst")).upper(),(data.get("gMiddle")).upper(),(data.get("gLast")).upper()
        app['B11'], app['N11'] =(data.get("gBday")), g_age
        app['B12'], app['L12'] =g_town_prov,(data.get("gCountry", "Philippines"))
        app['B13'], app['H13'] = "Male",(data.get("gCitizen", "Filipino"))
        app['B15'], app['M15'] = g_full_addr,(data.get("gCountry", "Philippines"))
        app['B16'], app['B17'] =(data.get("gReligion")),(data.get("gStatus", "Single"))
        
        # Parents & Givers (Groom)
        app['B22'], app['H22'], app['L22'] =(data.get("gFathF")),(data.get("gFathM")),(data.get("gFathL"))
        app['B26'], app['G26'], app['K26'] =(data.get("gMothF")),(data.get("gMothM")),(data.get("gMothL"))
        if 18 <= g_age <= 24:
            app['B30'], app['H30'], app['L30'] =(data.get("gGiverF")),(data.get("gGiverM")),(data.get("gGiverL"))
            app['B31'], app['B32'] =(data.get("gGiverRelation")),(data.get("gCitizen", "Filipino"))

        # Bride
        app['U8'], app['U9'], app['U10'] =(data.get("bFirst")).upper(),(data.get("bMiddle")).upper(),(data.get("bLast")).upper()
        app['U11'], app['AF11'] =(data.get("bBday")), b_age
        app['U12'], app['AE12'] =b_town_prov,(data.get("bCountry", "Philippines"))
        app['U13'], app['Z13'] = "Female",(data.get("bCitizen", "Filipino"))
        app['U15'], app['AF15'] = b_full_addr,(data.get("bCountry", "Philippines"))
        app['U16'], app['U17'] =(data.get("bReligion")),(data.get("bStatus", "Single"))

        # Parents & Givers (Bride)
        app['U22'], app['Y22'], app['AC22'] =(data.get("bFathF")),(data.get("bFathM")),(data.get("bFathL"))
        app['U26'], app['Y26'], app['AD26'] =(data.get("bMothF")),(data.get("bMothM")),(data.get("bMothL"))
        if 18 <= b_age <= 24:
            app['U30'], app['Y30'], app['AD30'] =(data.get("bGiverF")),(data.get("bGiverM")),  (data.get("bGiverL"))
            app['U31'], app['U32'] = (data.get("bGiverRelation")),(data.get("bCitizen", "Filipino"))

        # Common Footer Info
        app['B37'], app['U37'], app['E37'], app['W37'], app['L37'], app['AD37'] = day_now, day_now, month_now, month_now, year_now, year_now
        app['B38'] = app['U38'] = "Solano, Nueva Vizcaya"

        # Sheet Visibility Logic
        is_groom_ext = g_town_prov != "Solano, Nueva Vizcaya"
        is_bride_ext = b_town_prov != "Solano, Nueva Vizcaya"
        
        sheets_to_keep = ["APPLICATION", "Notice"]
        
        # Age-based Consent/Advice Sheets
        extra_sheet = None
        if 18 <= b_age <= 20 and g_age >= 25: extra_sheet = "CONSENT F"
        elif 18 <= g_age <= 20 and b_age >= 25: extra_sheet = "CONSENT M"
        elif 18 <= b_age <= 20 and 18 <= g_age <= 20: extra_sheet = "CONSENT M&F"
        elif 21 <= b_age <= 24 and g_age >= 25: extra_sheet = "ADVICE F"
        elif 21 <= g_age <= 24 and b_age >= 25: extra_sheet = "ADVICE M"
        elif 21 <= b_age <= 24 and 21 <= g_age <= 24: extra_sheet = "ADVICE M&F"
        elif 21 <= g_age <= 24 and 18 <= b_age <= 20: extra_sheet = "ADVICE M-CONSENT F"
        elif 21 <= b_age <= 24 and 18 <= g_age <= 20: extra_sheet = "ADVICE F-CONSENT M"
        
        if extra_sheet: sheets_to_keep.append(extra_sheet)
        if is_groom_ext or is_bride_ext: sheets_to_keep.extend(["AddressBACKnotice", "EnvelopeAddress"])

        # Delete unwanted sheets
        for s in wb.sheetnames:
            if s not in sheets_to_keep:
                del wb[s]

        # OUTPUT STREAMS
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        sys.stdout.buffer.write(output.read())
        sys.stdout.buffer.flush()
        os._exit(0)

    except Exception as e:
        sys.stderr.write(f"Error during processing: {str(e)}\n")
        os._exit(1)

if __name__ == "__main__":
    # Wait for input from stdin (Next.js sends data here)
    try:
        input_raw = sys.stdin.read()
        if not input_raw:
            sys.stderr.write("No data received\n")
            os._exit(1)
        
        input_json = json.loads(input_raw)
        process_excel(input_json)
    except Exception as e:
        sys.stderr.write(f"Startup Error: {str(e)}\n")
        os._exit(1)